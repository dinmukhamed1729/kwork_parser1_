import logging
import os
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter.ttk import Progressbar

import openpyxl

from parser import Parser


def get_base_url_and_directory(excel_path):
    if os.path.exists(excel_path):
        # Загружаем рабочую книгу
        workbook = openpyxl.load_workbook(excel_path)

        # Проверяем, есть ли скрытый лист config
        if "config" in workbook.sheetnames:
            sheet = workbook["config"]

            # Получаем глобальный URL из первой ячейки
            base_url = sheet.cell(row=1, column=1).value

            # Получаем путь к папке, где находится файл
            directory_path = os.path.dirname(excel_path)

            return base_url, directory_path
        else:
            print("Лист 'config' не найден.")
            return None, None
    else:
        print(f"Файл {excel_path} не найден.")
        return None, None


class UrlInputPage:
    def __init__(self, root, session):
        self.result_text = None
        self.path_entry = None
        self.url_entry = None
        self.progress = None
        self.root = root
        self.session = session
        self.directory = None
        self.save_path = tk.StringVar()
        self.report_path = tk.StringVar()

        self.parser = None

        self.create_ui()

    def create_ui(self):
        tk.Label(self.root, text="Введите URL форума:").grid(row=0, column=0, padx=10, pady=5)
        self.url_entry = tk.Entry(self.root, width=50)
        self.url_entry.grid(row=0, column=1, padx=10, pady=5)

        self.progress = Progressbar(self.root, orient="horizontal", length=300, mode="determinate")
        self.progress.grid(row=4, column=0, columnspan=3, pady=10)

        # Поле выбора каталога
        tk.Label(self.root, text="Каталог для скачивания:").grid(row=1, column=0, padx=10, pady=5)
        self.path_entry = tk.Entry(self.root, textvariable=self.save_path, width=40)
        self.path_entry.grid(row=1, column=1, padx=10, pady=5)
        tk.Button(self.root, text="Выбрать", command=self.select_directory).grid(row=1, column=2, padx=5, pady=5)

        # Поле для ввода пути к файлу report.xlsx
        tk.Label(self.root, text="Путь к файлу report.xlsx:").grid(row=2, column=0, padx=10, pady=5)
        self.report_entry = tk.Entry(self.root, textvariable=self.report_path, width=40)
        self.report_entry.grid(row=2, column=1, padx=10, pady=5)
        tk.Button(self.root, text="Выбрать", command=self.select_report_file).grid(row=2, column=2, padx=5, pady=5)

        # Кнопки управления
        tk.Button(self.root, text="Парсить форум", command=self.parse_forum).grid(row=3, column=0, columnspan=3,
                                                                                  pady=10)

        # Поле для вывода результатов
        self.result_text = tk.Text(self.root, height=10, width=70)
        self.result_text.grid(row=5, column=0, columnspan=3, padx=10, pady=10)

    def update_progress(self, value):
        self.progress["value"] = value
        self.root.update_idletasks()

    def select_directory(self):
        """Выбор каталога для сохранения скачанных файлов."""
        self.directory = filedialog.askdirectory()
        if self.directory:
            self.save_path.set(self.directory)  # Устанавливаем путь
            logging.info(f"Каталог сохранения установлен: {self.directory}")

    def select_report_file(self):
        """Выбор файла report.xlsx"""
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            self.report_path.set(file_path)  # Устанавливаем путь
            logging.info(f"Путь к файлу report.xlsx: {file_path}")

    def parse_forum(self):
        """Запускает парсинг форума."""
        url = self.url_entry.get().strip()
        save_path = self.save_path.get().strip()
        report_file_path = self.report_path.get().strip()

        if not url or not save_path:
            if not report_file_path:
                messagebox.showwarning("Ошибка", "Введите URL или выберите файл report.xlsx!")
                return

            # Получаем данные из report.xlsx
            base_url, directory = get_base_url_and_directory(report_file_path)

            if not base_url and not url:
                messagebox.showwarning("Ошибка", "Не удалось получить URL из report.xlsx!")
                return

            if not directory and not save_path:
                messagebox.showwarning("Ошибка", "Не удалось получить путь к папке из report.xlsx!")
                return

            # Подставляем данные, если их нет
            if not url:
                url = base_url
                self.url_entry.insert(0, url)  # Заполняем поле ввода

            if not save_path:
                save_path = directory
                self.save_path.set(save_path)  # Обновляем поле пути

        # Создаём объект парсера
        self.parser = Parser(self.session, url, save_path)
        print(f"Запуск парсинга форума: {url}")

        results = self.parser.parse_forum()

        if results:
            self.result_text.delete("1.0", tk.END)
            for thread in results:
                self.result_text.insert(tk.END, f"{thread['title']} - {thread['author']}\n")

            messagebox.showinfo("Успех", f"Найдено {len(results)} тем.")
        else:
            messagebox.showwarning("Ошибка", "Темы не найдены.")
