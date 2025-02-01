import os
import re
import logging
from urllib.parse import urlsplit, urljoin, unquote

import openpyxl
import pandas as pd
import requests
from bs4 import BeautifulSoup


class Parser:
    def __init__(self, session, main_url, save_path="./"):
        """Конструктор парсера."""
        self.session = session
        self.save_path = save_path
        self.main_url = main_url
        self.base_url = "https://ecu-firmware-files.ru"
        self.excel_path = os.path.join(self.save_path, "report.xlsx")
        if not os.path.exists(save_path):
            os.makedirs(save_path)

    def get_page_content(self, url):
        """Запрашивает HTML-код страницы с обработкой ошибок."""
        try:
            response = self.session.get(url, timeout=10)
            response.raise_for_status()
            return response.text
        except requests.exceptions.Timeout:
            logging.error(f"⏳ Таймаут при загрузке {url}")
        except requests.exceptions.RequestException as e:
            logging.error(f"❌ Ошибка сети {url}: {e}")
        return None

    def find_thread_links(self, html):
        """Находит ссылки на темы."""
        soup = BeautifulSoup(html, "html.parser")
        return [
            self.base_url + a["href"].split("/unread")[0]
            for div in soup.find_all("div", class_="structItem-title")
            if (a := div.find("a", href=True))
        ]

    def extract_thread_id(self, thread_url):
        """Извлекает ID темы из URL."""
        match = re.search(r"/threads/([^/]+)(?:/|$)", thread_url)
        return match.group(1) if match else None

    def extract_articles(self, html):
        """Извлекает статьи из темы."""

        soup = BeautifulSoup(html, "html.parser")

        container = soup.find("div", class_="block-body js-replyNewMessageContainer")

        if not container:
            return []

        return container.find_all("article", class_="message")

    def parse_thread(self, thread_url):
        """Парсит страницу темы и извлекает данные."""
        html = self.get_page_content(thread_url)
        if not html:
            return None
        articles = self.extract_articles(html)

        titles, authors, texts = [], [], []

        for article in articles:
            author = article.find("a", class_="username")
            description = article.find("div", class_="bbCodeBlock-expandContent") or article.find("div",
                                                                                                  class_="bbWrapper")

            author = author.get_text(strip=True) if author else "Автор не найден"
            text = description.get_text("\n", strip=True) if description else "Текст не найден"

            authors.append(author)
            texts.append(text)

        # Сбор ссылок на вложенные файлы

        soup = BeautifulSoup(html, "html.parser")
        attachments = [
            {"name": link["title"], "url": link["href"]}
            for link in soup.select("ul.attachmentList a[href]")
            if link.get("title")
        ]

        # Поиск кнопки "Скачать"
        download_button = soup.select_one(".p-title-pageAction a.button--cta")
        if download_button and (download_url := download_button.get("href")):
            attachments.append({"name": "", "url": download_url})

        new_attachments = []
        for attachment in attachments:
            global_file_url = urljoin(self.base_url, attachment["url"])
            response = self.session.get(global_file_url, stream=True)
            response.raise_for_status()

            # Проверка Content-Type и Content-Disposition для различия файлов и HTML-страниц
            content_type = response.headers.get("Content-Type", "")

            if "text/html" in content_type:
                file_info = self.get_file_info(global_file_url)
                for file in file_info:
                    new_attachments.append({"name": file["name"], "url": file["url"]})
            else:
                new_attachments.append({"name": attachment["name"], "url": attachment["url"]})

        attachments = new_attachments

        title_tag = soup.select_one(".p-title .p-title-value")
        title = title_tag.text.strip() if title_tag else ""

        dir_name = re.sub(r'[\\/|?&"<>* :]', '_', title)

        return {
            "title": title,
            "author": authors,
            "br_text": texts,
            "attachments": attachments,
            "dir_name": dir_name,
            "thread_url": thread_url,
        }

    def parse_forum(self):
        """Парсит все страницы форума и сохраняет данные."""
        page_number = 1
        results = []
        forum_url = self.main_url
        while forum_url:
            html = self.get_page_content(forum_url)
            if not html:
                logging.error(f"Не удалось загрузить {forum_url}.")
                break

            thread_links = self.find_thread_links(html)
            for thread_url in thread_links:
                if thread_data := self.parse_thread(thread_url):
                    results.append(thread_data)
                    self.save_thread_data(thread_data)

            # Переход на следующую страницу
            soup = BeautifulSoup(html, "html.parser")
            next_page = soup.select_one("a.pageNav-jump--next")
            forum_url = self.base_url + next_page["href"] if next_page else None
            page_number += 1

        print(f"Парсинг завершен. Обработано {page_number - 1} страниц, найдено {len(results)} тем.")
        return results

    def save_thread_data(self, data):
        """Сохраняет данные темы: файлы, текст, отчет в Excel."""
        thread_folder = os.path.join(self.save_path, data["dir_name"])
        os.makedirs(thread_folder, exist_ok=True)

        # 1️⃣ Сохранение текстового файла с описанием
        self.save_text_file(data, thread_folder)

        # 2️⃣ Поиск и скачивание вложенных файлов
        self.download_attachments(data, thread_folder)

        # 3️⃣ Добавление в `report.xlsx`
        self.update_report(data)

    def save_text_file(self, data, thread_folder):
        """Сохраняет текстовый файл с описанием темы, избегая дубликатов."""
        txt_path = os.path.join(thread_folder, f"{data['dir_name']}.txt")

        # Нормализация ссылки на тему (убираем завершающий слеш)
        normalized_url = data['thread_url'].rstrip('/')

        new_entries = []

        # Формируем новые записи и объединяем их в единую строку
        for author, text in zip(data["author"], data["br_text"]):
            normalized_text = text.strip()  # Убираем лишние пробелы и пустые строки
            entry = f"Ссылка на тему: {normalized_url}\nНазвание: {data['title']}\nАвтор: {author}\nОписание:\n{normalized_text}\n"
            new_entries.append(entry)

        new_content = "".join(new_entries)  # Превращаем новые записи в единую строку

        # Читаем весь файл в строку
        existing_content = ""
        if os.path.exists(txt_path):
            with open(txt_path, "r", encoding="utf-8") as f:
                existing_content = f.read()

        if new_content not in existing_content:
            with open(txt_path, "a", encoding="utf-8") as f:
                f.write(new_content)
                f.write("\n" * 5)
            print(f"✅ Текстовый файл с описанием сохранен: {txt_path}")
        else:
            print(f"⚠️ Запись уже существует, файл пропущен: {txt_path}")

    def download_attachments(self, data, thread_folder):
        """Загружает все вложенные файлы в тему."""
        for attachment in data["attachments"]:
            global_file_url = urljoin(self.base_url, attachment["url"])
            file_name = attachment["name"]

            try:
                if not self.check_file_url_exists(attachment["url"]):
                    self.download_file(global_file_url, thread_folder, file_name)
                else:
                    print(attachment["name"],"пропущен")
            except requests.RequestException as e:
                logging.error(f"❌ Ошибка скачивания {global_file_url}: {e}")

    def download_file(self, global_file_url, thread_folder, file_name=""):
        """Скачивает файл по ссылке и сохраняет его на диск."""

        response = self.session.get(global_file_url, stream=True)
        response.raise_for_status()

        # Получаем корректное имя файла
        file_name = self.get_filename_from_headers(response.headers, global_file_url, file_name)

        if file_name == "reply":
            print(f"⚠️ Пропущен файл с именем 'reply': {global_file_url}")
            return

        file_path = os.path.join(thread_folder, file_name)
        file_path = self.ensure_unique_file_path(file_path)

        with open(file_path, "wb") as f:
            for chunk in response.iter_content(1024):
                f.write(chunk)

        print(f"✅ Файл {file_name} скачан и сохранен в {file_path}")

    def get_filename_from_headers(self, headers, file_url, file_name):
        """Извлекает имя файла из заголовков или URL, поддерживая кириллицу."""
        content_disposition = headers.get("Content-Disposition")

        if not file_name:
            # Проверяем `filename*` с UTF-8 кодировкой
            if content_disposition and "filename*" in content_disposition:
                file_name_encoded = content_disposition.split("filename*=")[1]
                encoding, file_name = file_name_encoded.split("''", 1)
                file_name = unquote(file_name)

            # Проверяем `filename`
            elif content_disposition and "filename=" in content_disposition:
                file_name = content_disposition.split("filename=")[1].strip('"')

            # Если нет заголовка, берем имя из URL
            else:
                file_name = os.path.basename(urlsplit(file_url).path)
                file_name = unquote(file_name)  # Декодируем кириллицу

        return file_name

    def ensure_unique_file_path(self, file_path):
        """Проверяет наличие файла с таким же именем и добавляет номер, если файл уже существует."""
        counter = 1
        original_file_path = file_path
        while os.path.exists(file_path):
            name, ext = os.path.splitext(original_file_path)
            file_path = os.path.join(os.path.dirname(original_file_path), f"{name} - {counter}{ext}")
            counter += 1
        return file_path

    def check_file_url_exists(self, file_url):
        """Метод для проверки, существует ли ссылка на файл в отчете."""
        if os.path.exists(self.excel_path):
            df = pd.read_excel(self.excel_path)
        else:
            df = pd.DataFrame(
                columns=["№", "Статус", "Название темы", "Ссылка на тему", "Ссылка на файл", "Название файла"])

        # Проверяем, существует ли file_url в колонке "Ссылка на файл"
        if file_url in df["Ссылка на файл"].values:
            print(f"Ссылка на файл {file_url} уже существует в отчете.")
            return True
        else:
            return False

    def update_report(self, data):
        """Обновляет отчет в Excel."""

        df = pd.read_excel(self.excel_path) if os.path.exists(self.excel_path) else pd.DataFrame(
            columns=["№", "Статус", "Название темы", "Ссылка на тему", "Ссылка на файл", "Название файла"]
        )
        for attachment in data["attachments"]:
            if not self.check_file_url_exists(attachment["url"]):
                df = pd.concat([df, pd.DataFrame([{
                    "№": len(df) + 1,
                    "Статус": "Скачан",
                    "Название темы": data["title"],
                    "Ссылка на тему": data["thread_url"],
                    "Ссылка на файл": attachment["url"],
                    "Название файла": attachment["name"]
                }])], ignore_index=True)
                print(f"📊 Отчет обновлен: {self.excel_path}")

        df.to_excel(self.excel_path, index=False)

        self.check_and_add_base_url()

    def check_and_add_base_url(self):
        config_sheet_name = "config"

        if os.path.exists(self.excel_path):
            workbook = openpyxl.load_workbook(self.excel_path)
            # Проверяем, существует ли лист config, если нет - создаем его
            if config_sheet_name not in workbook.sheetnames:
                sheet = workbook.create_sheet(config_sheet_name)
                # Записываем глобальный URL в первую ячейку
                sheet.cell(row=1, column=1).value = self.main_url
                sheet.sheet_state = "hidden"  # Скрываем лист
            else:
                sheet = workbook[config_sheet_name]
                # Если URL не записан, добавляем его
                if not sheet.cell(row=1, column=1).value:
                    sheet.cell(row=1, column=1).value = self.main_url
                    sheet.sheet_state = "hidden"  # Скрываем лист

            workbook.save(self.excel_path)
        else:
            print("Файл отчета не найден!")

    def get_file_info(self, page_url):
        """Извлекает имя файла и ссылку для скачивания с HTML страницы."""
        try:
            response = self.session.get(page_url)
            response.raise_for_status()

            soup = BeautifulSoup(response.text, "html.parser")
            file_info = []

            file_items = soup.select(".block-body .block-row")

            for item in file_items:
                file_name_tag = item.select_one(".contentRow-title")
                file_link_tag = item.select_one(".contentRow-extra a")

                if file_name_tag and file_link_tag:
                    file_name = file_name_tag.text.strip()
                    file_url = urljoin(self.base_url, file_link_tag["href"])

                    file_info.append({"name": file_name, "url": file_url})

            return file_info

        except requests.RequestException as e:
            print(f"Ошибка при получении страницы: {e}")
            return []
